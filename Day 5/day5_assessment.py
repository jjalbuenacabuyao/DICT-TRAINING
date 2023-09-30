from tkinter import *
from tkinter import ttk, messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

FONT = "SegoeUI 14"
BG_COLOR = "#242424"

root = Tk()
root.geometry("1300x600")
root.configure(bg=BG_COLOR)
root.title("CRUD IN TKINTER")


excel_con = Workbook()
excel_con = load_workbook("register.xlsx")
excel_opr = excel_con.active


namelbl = Label(root, text="Name", font=FONT)
namelbl.grid(
    row=0, 
    column=0, 
    sticky=W
)

name_entry = Entry(
    root, 
    width=40, 
    font=FONT
)

name_entry.grid(
    row=0, 
    column=1, 
    pady=5, 
    sticky=W
)

programlbl = Label(
    root, 
    text="Program Enrolled", 
    font=FONT
)

programlbl.grid(
    row=1, 
    column=0, 
    pady=5, 
    sticky=W
)

program_selected = StringVar()

BSIT = Radiobutton(
    root, 
    text="BS Information Technology", 
    variable=program_selected, 
    value="BSIT"
)

BSCS = Radiobutton(
    root, 
    text="BS Computer Science", 
    variable=program_selected, 
    value="BSCS"
)

BSCE = Radiobutton(
    root, 
    text="BS Computer Engineering", 
    variable=program_selected, 
    value="BSCE"
)

BSIS = Radiobutton(
    root, 
    text="BS Information System", 
    variable=program_selected, 
    value="BSIS"
)

BSIT.grid(
    row=2, 
    column=1, 
    pady=2, 
    sticky=W
)

BSCS.grid(
    row=3, 
    column=1, 
    pady=2, 
    sticky=W
)

BSCE.grid(
    row=4, 
    column=1, 
    pady=2, 
    sticky=W
)

BSIS.grid(
    row=5, 
    column=1, 
    pady=2, 
    sticky=W
)

addlbl = Label(
    root, 
    text="Residence", 
    font=FONT
)

addlbl.grid(
    row=6, 
    column=0, 
    pady=3, 
    sticky=W
)

address = Text(
    root, 
    width=20, 
    height=4
)

address.grid(
    row=6, 
    column=1, 
    pady=8, 
    sticky=W
)

submit = Button(
    root, 
    text="Submit record", 
    width=12, 
    height=2, 
    font=FONT, 
    command=lambda:submit_data()
)

submit.grid(
    row=7, 
    column=1, 
    sticky=W
)

search = Button(
    root, 
    text="Search", 
    width=12, 
    height=2, 
    font=FONT, 
    command=lambda:search()
)
search.grid(
    row=8, 
    column=1, 
    sticky=W, 
    pady=10
)

refresh = Button(
    root, 
    text="Refresh", 
    width=12, 
    height=2, 
    font=FONT, 
    command=lambda:refresh_tree()
)
refresh.grid(
    row=9, 
    column=1, 
    sticky=W
)

delete = Button(
    root, 
    text="Delete", 
    width=12, 
    height=2, 
    font=FONT, 
    command=lambda:delete_data()
)
delete.grid(
    row=10, 
    column=1, 
    sticky=W
)

tree = ttk.Treeview(
    root, 
    columns=(
        "Names", 
        "Program", 
        "Address"
    ), 
    show="headings"
)
tree.heading("Names", text="Names")
tree.heading("Program", text="Program")
tree.heading("Address", text="Address")

tree.grid(
    row=0, 
    column=3, 
    rowspan=10, 
    ipadx=8
)


def submit_data():
    global nametxt
    global programtxt
    global addresstxt

    nametxt = name_entry.get()
    programtxt = program_selected.get()
    addresstxt = address.get("1.0", "end-1c")

    last_row = str(excel_opr.max_row + 1)

    excel_opr["A" + last_row] = nametxt
    excel_opr["B" + last_row] = programtxt
    excel_opr["C" + last_row] = addresstxt

    excel_con.save("register.xlsx")
    message = f"Data added successfully"
    messagebox.showinfo("RECORD", message)

    tree.insert(
            "", 
            "end", 
            values=(
                nametxt,
                programtxt,
                addresstxt
            )
        )



def view_data():
    for cell in range(1, excel_opr.max_row + 1):
        tree.insert(
            "", 
            "end", 
            values=(
                excel_opr["A" + str(cell)].value, 
                excel_opr["B" + str(cell)].value, 
                excel_opr["C" + str(cell)].value
            )
        )


def search():
    global cell_address
    nametxt = name_entry.get()
    found = False

    for row in range(1, excel_opr.max_row + 1):
        if nametxt == excel_opr["A" + str(row)].value:
            found = True
            cell_address = str(row)
            break

    if found:
        name_entry.insert(0, excel_opr["A" + cell_address].value)
        program_value = excel_opr["B"+cell_address].value

        if program_value == "BSIT":
            BSIT.select()
        elif program_value == "BSCS":
            BSCS.select()
        elif program_value == "BSCE":
            BSCE.select()
        else:
            BSIS.select()

        address.insert("1.0", excel_opr["C" + cell_address].value)
    

def delete_data():
    excel_opr.delete_rows(int(cell_address))
    messagebox.showerror("Data deleted", "Data removed from database")
    excel_con.save("register.xlsx")


def get_updated_data():
    updated_values = list()
    for row in range(1, excel_opr.max_row + 1):
        updated_values.append(
            [excel_opr["A" + str(row)].value, excel_opr["B" + str(row)].value, excel_opr["C" + str(row)].value]
            )
    
    return updated_values


def refresh_tree():
    tree.delete(*tree.get_children())
    data = get_updated_data()
    for i in data:
        tree.insert("", "end", values=i)

view_data()

root.mainloop()
